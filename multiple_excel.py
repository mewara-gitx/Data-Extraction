import os
import pandas as pd
from openpyxl import load_workbook
import openpyxl

#finding the files by name in the particular Drive

def find_files():
    print("Enter the file name : ")
    filename = input()
    result = []
    for root, dir, files in os.walk("D:"):
        if filename in files:
            result.append(os.path.join(root, filename))
    return result

# add another file in the same drive but differnt location 
def file_finder():
    searched_data=find_files()
    #save_data_mastersheet(searched_data)
    print('Wanna search more?????...')
    ans_more=input()
    while ans_more.lower() == 'yes':
        more_searched_data = []
        more_searched_data=find_files()
        searched_data.extend(more_searched_data)
        print('Wanna search more?????...')
        ans_more=input()
        if ans_more.lower()=='no':
            break

    return searched_data


# search function to search the particular number by help of unique ID
def search_id(df, ps_number):
    result = df[df['PS number'] == ps_number]
    return result  

# get the sheet Data from all excel books in sheet 1
def sheet_data():
    files_xls = file_finder()
    sheets ={}
    for i in range(len(files_xls)):
        sheets['from '+ str(i)] = pd.read_excel(files_xls[i], 'Sheet1')
    return sheets

# if data is found in the sheets then all rows of the matching unique ID is taken
def match_unique():
    dict_sheets = sheet_data()                        # the function sheet access is assigned to the new variable
    results={}
    print("Enter the Unique ID you want to search : ")
    num = int(input())
    for j in range(len(dict_sheets.keys())):
        results['from_sheet'+str(j)]=search_id(dict_sheets['from '+str(j)],num)

    if len(dict_sheets.keys()) > 1:
        all_data = pd.merge(results['from_sheet0'], results['from_sheet1'], how='left')
        for j in range(2, len(dict_sheets.keys())):
            all_data = pd.merge(all_data, results['from_sheet' + str(j)], how='left')
    else:
        all_data = results['from_sheet0']

    return all_data


#this function saves to new excel work book
def save_to_excel():
    if not os.path.isdir("D:"):
        os.makedirs("D:")


    if not os.path.isfile('FileName.xlsx'):
        wb = openpyxl.Workbook()  
        dest_filename = 'FileName.xlsx' 
        path = os.path.join('D:', dest_filename)
        wb.save(path)
    final = match_unique()
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    if 'Sheet' in book.sheetnames:
        pfd = book['Sheet']
        book.remove(pfd)
    final.to_excel(writer, sheet_name = 'Sheet')
    writer.save()
    writer.close()
    
save_to_excel()
    