"""
Importing necessary modules
"""

import pandas as pd
from openpyxl import load_workbook

"""
Function to take user input how many sheets to scan 
or to scan all the sheets present in the excel file
and store the result in dictionary(named sheets)
"""


def sheets_access():
    workbook = load_workbook(filename="Book1.xlsx")
    res = workbook.sheetnames
    print("Total number of sheets in excel file are " + str(len(res)))
    print("Do you want to search data in all sheets: ")
    answer = input()
    sheets = {}                                          # empty dictionary
    if(answer.lower() == 'yes'):
        for i in range(len(res)):                        # for loop for multiple sheets in the excelbook
            sheets["Sheet" + str(i)] = pd.read_excel("Book1.xlsx",engine = "openpyxl", sheet_name=i)
            sheets["Sheet" + str(i)].dropna(axis=1, how='all', inplace=True)
#             print(sheets["Sheet" + str(i)])
    else:
        print("Enter the number of sheets you want to scan")
        num = int(input())
        for i in range(num):
            sheets["Sheet" + str(i)] = pd.read_excel("Book1.xlsx",engine = "openpyxl",sheet_name=i)
            sheets["Sheet" + str(i)].dropna(axis=1, how='all',inplace=True)
#             print(sheets["Sheet" + str(i)])
    return sheets

"""
Function to search by unique ID in all the excel sheet  
"""

def search_id(df, ps_number):
    result = df[df['PS number'] == ps_number]
    return result                                     # result is the dataframe of all the matching information of the unique ID

"""
Function to match unique ID in sheets present in the excel book
"""
def match_unique(dict_sheets):          # the function sheet access is assigned to the new variable
    results={}
    print("Enter the Unique ID you want to search : ")
    num = int(input())
    for j in range(len(dict_sheets.keys())):
        results['from_sheet'+str(j)]=search_id(dict_sheets['Sheet'+str(j)],num)

    if len(dict_sheets.keys()) > 1:
        all_data = pd.merge(results['from_sheet0'], results['from_sheet1'], how='left')
        for j in range(2, len(dict_sheets.keys())):
            all_data = pd.merge(all_data, results['from_sheet' + str(j)], how='left')
    else:
        all_data = results['from_sheet0']

    return all_data

# this function is to save data frame of matching unique ID int the mastersheet
def save_data_mastersheet(final):
    #final = match_unique()
    path = r"Book1.xlsx"
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    if 'mastersheet' in book.sheetnames:
        pfd = book['mastersheet']
        book.remove(pfd)
    final.to_excel(writer, sheet_name='mastersheet')

    writer.save()
    writer.close()

# this is the main function to search multiple data from the sheets 
if __name__=="__main__":
    dict_of_sheets=sheets_access()
    searched_data=match_unique(dict_of_sheets)
    print(searched_data)
    #save_data_mastersheet(searched_data)
    print('Wanna search more?????...')
    ans_more=input()
    while ans_more.lower()== 'yes':
        more_searched_data=match_unique(dict_of_sheets)
        searched_data=searched_data.append(more_searched_data, ignore_index = True) 
        print('Wanna search more?????...')
        ans_more=input()
        if ans_more.lower()=='no':
            break

    print(searched_data)
    save_data_mastersheet(searched_data)    